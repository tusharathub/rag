import os
from openpyxl import load_workbook
from app.interfaces.parsers import (
    IDocumentParser, 
    ParsedDocumentResult, 
    ParsedPage, 
    TableData, 
    DocumentParsingError
)


class ExcelParser(IDocumentParser):
    """Parser implementation for Excel spreadsheets (XLSX, XLS)."""

    def parse(self, file_path: str) -> ParsedDocumentResult:
        if not os.path.exists(file_path):
            raise DocumentParsingError(f"File not found: {file_path}")
            
        try:
            # load_workbook in read-only and data_only mode for memory safety and formula evaluation
            wb = load_workbook(file_path, data_only=True, read_only=True)
            pages = []
            all_text_parts = []
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                page_number = sheet_idx + 1
                
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out fully empty rows
                    if all(cell is None for cell in row):
                        continue
                    row_cells = [str(cell) if cell is not None else "" for cell in row]
                    rows.append(row_cells)
                
                tables = []
                text_parts = []
                
                if rows:
                    headers = rows[0]
                    table_rows = rows[1:] if len(rows) > 1 else []
                    
                    tables.append(TableData(
                        headers=headers,
                        rows=table_rows,
                        caption=f"Worksheet: {sheet_name}"
                    ))
                    
                    text_parts.append(f"Worksheet: {sheet_name}")
                    text_parts.append(" | ".join(headers))
                    for tr in table_rows:
                        text_parts.append(" | ".join(tr))
                
                sheet_text = "\n".join(text_parts)
                pages.append(ParsedPage(
                    page_number=page_number,
                    text=sheet_text,
                    headings=[],
                    tables=tables,
                    images=[]  # Images are not extractable in openpyxl read_only mode
                ))
                
                if sheet_text:
                    all_text_parts.append(sheet_text)
            
            wb.close()
            
            raw_text = "\n\n".join(all_text_parts)
            metadata = {
                "sheet_count": len(wb.sheetnames),
                "sheet_names": ", ".join(wb.sheetnames)
            }
            
            return ParsedDocumentResult(
                raw_text=raw_text,
                metadata=metadata,
                pages=pages
            )
        except Exception as e:
            raise DocumentParsingError(f"Failed parsing Excel file: {str(e)}")
